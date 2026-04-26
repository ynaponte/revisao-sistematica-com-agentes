"""Spreadsheet I/O: read articles from .csv, .xls, and .xlsx, and write results to .xlsx."""

from __future__ import annotations

from pathlib import Path
from typing import Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import Article


def parse_row_range(row_range_str: str, total_data_rows: int) -> tuple[int, int]:
    parts = row_range_str.strip().split("-")
    if len(parts) == 1:
        start = int(parts[0]) - 1
        return start, start
    start = int(parts[0]) - 1
    end = int(parts[1]) - 1
    end = min(end, total_data_rows - 1)
    return max(0, start), max(0, end)


def _find_metadata(df: pd.DataFrame) -> dict | None:
    """Scans the dataframe heuristically to find the header row and column indices."""
    for row_idx in range(min(50, len(df))):
        row = df.iloc[row_idx]
        title_col_idx = None
        abstract_col_idx = None
        id_col_idx = None
        
        for col_idx, cell_value in enumerate(row):
            val = str(cell_value).lower().strip()
            # Common matches for Title
            if not title_col_idx and ("title" in val or "título" in val or "titulo" in val):
                title_col_idx = col_idx
            # Common matches for Abstract
            elif not abstract_col_idx and ("abstract" in val or "resumo" in val):
                abstract_col_idx = col_idx
            # Common matches for ID
            elif not id_col_idx and val in ["id", "identificador", "key", "nº", "no.", "n"]:
                id_col_idx = col_idx
                
        if title_col_idx is not None and abstract_col_idx is not None:
            return {
                "header_row": row_idx,
                "title_col": title_col_idx,
                "abstract_col": abstract_col_idx,
                "id_col": id_col_idx
            }
    return None


def load_articles(
    path: str | Path,
    sheet_name: str | None = None,
    row_range: str | None = None,
) -> list[Article]:
    path_obj = Path(path)
    path_str = str(path_obj).lower()
    dfs = {}
    
    try:
        if path_str.endswith('.csv'):
            dfs['csv_data'] = pd.read_csv(path_obj, header=None)
        elif path_str.endswith('.xls'):
            dfs = pd.read_excel(path_obj, sheet_name=None, header=None, engine="xlrd")
        elif path_str.endswith('.xlsx'):
            dfs = pd.read_excel(path_obj, sheet_name=None, header=None, engine="openpyxl")
        else:
            # Fallback attempt
            try:
                dfs = pd.read_excel(path_obj, sheet_name=None, header=None)
            except Exception:
                dfs['csv_data'] = pd.read_csv(path_obj, header=None)
    except Exception as e:
        print(f"Failed to load file {path}: {e}")
        return []

    selected_df = None
    meta = None
    
    # Heuristic search across all sheets to find the first valid one
    for s_name, df in dfs.items():
        if df.empty:
            continue
        meta = _find_metadata(df)
        if meta:
            selected_df = df
            break
            
    if selected_df is None or meta is None:
        return []
        
    start_data_row = meta["header_row"] + 1
    total_data_rows = len(selected_df) - start_data_row
    
    if total_data_rows <= 0:
        return []
        
    if row_range:
        range_start, range_end = parse_row_range(row_range, total_data_rows)
    else:
        range_start, range_end = 0, total_data_rows - 1

    articles: list[Article] = []
    
    for i in range(range_start, range_end + 1):
        row_idx = start_data_row + i
        if row_idx >= len(selected_df):
            break
            
        row_data = selected_df.iloc[row_idx]
        
        # Extract ID
        if meta["id_col"] is not None:
            raw_id = row_data[meta["id_col"]]
            if pd.isna(raw_id):
                raw_id = i + 1
        else:
            raw_id = i + 1
            
        try:
            # Handle cases where ID is a float like "1.0" or string
            art_id = int(float(raw_id)) if str(raw_id).replace('.','',1).isdigit() else i + 1
        except ValueError:
            art_id = i + 1
            
        title = str(row_data[meta["title_col"]]).strip()
        abstract = str(row_data[meta["abstract_col"]]).strip()
        
        # Skip empty rows where title is missing
        if not title or title.lower() == "nan":
            continue

        articles.append(
            Article(
                id=art_id,
                title=title,
                abstract=abstract if (abstract and abstract.lower() != "nan") else "(no abstract available)",
                row_index=row_idx,
            )
        )

    return articles


def write_results(
    articles: Sequence[Article],
    results: Sequence[dict], 
    inclusion_criteria: Sequence[str],
    exclusion_criteria: Sequence[str],
    output_path: str | Path,
    metadata: dict[str, str] | None = None,
) -> Path:
    """Export results to an Excel file using Pandas, with styled rows."""
    output_path = Path(output_path)
    
    data = []
    for article, result in zip(articles, results):
        data.append({
            "ID": article.id,
            "Title": article.title,
            "Abstract": article.abstract,
            "Decision": result.get("decision", "ERROR"),
            "Discriminants": ", ".join(result.get("rejection_reasons", [])),
            "Justification": result.get("justification", ""),
        })
        
    df = pd.DataFrame(data)
    
    # Create the writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        start_row = 0
        if metadata:
            meta_df = pd.DataFrame([metadata])
            meta_df.to_excel(writer, sheet_name="Screening Results", index=False, startrow=0)
            start_row = 3
            
        df.to_excel(writer, sheet_name="Screening Results", index=False, startrow=start_row)

        ws = writer.sheets["Screening Results"]
        
        # Adjust column widths
        widths = {"A": 8, "B": 50, "C": 60, "D": 12, "E": 25, "F": 80}
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

        # Apply basic text wrap to all cells
        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = wrap_alignment
                
        # Color specific decisions
        decision_col_idx = 4 # 1-based index for 'Decision' column
        for row in ws.iter_rows(min_row=start_row + 2, max_row=ws.max_row):
            decision_cell = row[decision_col_idx - 1]
            if decision_cell.value == "ACCEPTED":
                decision_cell.fill = PatternFill("solid", fgColor="C8E6C9")
            elif decision_cell.value == "REJECTED":
                decision_cell.fill = PatternFill("solid", fgColor="FFCDD2")

    return output_path
